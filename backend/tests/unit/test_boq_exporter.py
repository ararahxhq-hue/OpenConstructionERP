"""Unit tests for the BOQ (Bill of Quantities) workbook builder.

Structural / value tests rather than a byte snapshot - openpyxl embeds
build metadata that makes byte-level reproducibility brittle, so we pin
the timestamp via ``BoqExportOptions.frozen_now`` and assert on VALUES.
"""

from __future__ import annotations

import uuid
from datetime import UTC, datetime
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.modules.bim_hub.exporters import BoqExportOptions, build_boq_workbook
from app.modules.bim_hub.exporters.boq_xlsx import (
    DATA_START_ROW,
    DETAIL_COLUMNS,
    HEADER_ROW,
)


def _model() -> SimpleNamespace:
    return SimpleNamespace(
        id=uuid.UUID("11111111-1111-1111-1111-111111111111"),
        name="Tower A",
    )


def _elements() -> list[SimpleNamespace]:
    """Three IfcWall (one with alias keys, one with a bool that must be
    rejected), one IfcSlab (with a string-number weight), one IfcBeam."""
    return [
        SimpleNamespace(
            id=uuid.uuid4(),
            stable_id="wall-a",
            element_type="IfcWall",
            name="Wall A",
            storey="L1",
            discipline="Architectural",
            quantities={"area_m2": 10.0, "volume_m3": 2.0},
        ),
        SimpleNamespace(
            id=uuid.uuid4(),
            stable_id="wall-b",
            element_type="IfcWall",
            name="Wall B",
            storey="L2",
            discipline="Architectural",
            # alias keys: raw IFC BaseQuantity names must land in the
            # same columns as the canonical *_m2 / *_m3 keys.
            quantities={"area": 5.0, "NetVolume": 1.0},
        ),
        SimpleNamespace(
            id=uuid.uuid4(),
            stable_id="wall-bad",
            element_type="IfcWall",
            name="Wall Bad",
            storey="L1",
            discipline="Architectural",
            # a bool is an int subclass but must NOT be summed as 1.0
            quantities={"area_m2": True},
        ),
        SimpleNamespace(
            id=uuid.uuid4(),
            stable_id="slab-1",
            element_type="IfcSlab",
            name="Slab 1",
            storey="L1",
            discipline="Structural",
            quantities={"area_m2": 50.0, "volume_m3": 12.5, "weight_kg": "300"},
        ),
        SimpleNamespace(
            id=uuid.uuid4(),
            stable_id="beam-1",
            element_type="IfcBeam",
            name="Beam 1",
            storey="L1",
            discipline="Structural",
            quantities={"length_m": 8.0},
        ),
    ]


def _load(elements, **opts) -> object:
    blob = build_boq_workbook(
        _model(),
        elements,
        BoqExportOptions(frozen_now=datetime(2026, 6, 29, tzinfo=UTC), **opts),
    )
    return load_workbook(BytesIO(blob))


def _row(ws, r: int) -> list:
    return [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]


def test_sheets_present() -> None:
    wb = _load(_elements())
    assert wb.sheetnames == ["BOQ", "Elements"]


def test_include_detail_false_drops_elements_sheet() -> None:
    wb = _load(_elements(), include_detail=False)
    assert wb.sheetnames == ["BOQ"]


def test_summary_header_and_grouping_by_element_type() -> None:
    wb = _load(_elements())
    ws = wb["BOQ"]
    header = _row(ws, HEADER_ROW)
    assert header == [
        "#",
        "Element Type",
        "Count",
        "Area (m2)",
        "Volume (m3)",
        "Length (m)",
        "Weight (kg)",
    ]
    # Groups are alphabetical: IfcBeam, IfcSlab, IfcWall.
    beam = _row(ws, DATA_START_ROW)
    slab = _row(ws, DATA_START_ROW + 1)
    wall = _row(ws, DATA_START_ROW + 2)
    assert beam[1] == "IfcBeam"
    assert beam[2] == 1 and beam[5] == 8.0  # count, length
    assert slab[1] == "IfcSlab"
    assert slab[2] == 1 and slab[3] == 50.0 and slab[4] == 12.5 and slab[6] == 300.0
    assert wall[1] == "IfcWall"
    # 3 walls; alias keys summed; bool area rejected -> 10 + 5 + 0 = 15 area, 2 + 1 = 3 volume
    assert wall[2] == 3 and wall[3] == 15.0 and wall[4] == 3.0


def test_total_row() -> None:
    wb = _load(_elements())
    ws = wb["BOQ"]
    total = _row(ws, DATA_START_ROW + 3)
    assert total[0] == "TOTAL"
    assert total[2] == 5  # count
    assert total[3] == 65.0  # area = 15 + 50
    assert total[4] == 15.5  # volume = 3 + 12.5
    assert total[5] == 8.0  # length
    assert total[6] == 300.0  # weight


def test_group_by_element_type_storey() -> None:
    wb = _load(_elements(), group_by="element_type_storey")
    ws = wb["BOQ"]
    header = _row(ws, HEADER_ROW)
    assert header[:3] == ["#", "Element Type", "Storey"]
    # (IfcBeam,L1), (IfcSlab,L1), (IfcWall,L1)=2, (IfcWall,L2)=1
    rows = [_row(ws, DATA_START_ROW + i) for i in range(4)]
    keys = [(r[1], r[2], r[3]) for r in rows]  # type, storey, count
    assert keys == [
        ("IfcBeam", "L1", 1),
        ("IfcSlab", "L1", 1),
        ("IfcWall", "L1", 2),
        ("IfcWall", "L2", 1),
    ]


def test_group_by_storey() -> None:
    wb = _load(_elements(), group_by="storey")
    ws = wb["BOQ"]
    assert _row(ws, HEADER_ROW)[1] == "Storey"
    # L1 has beam+slab+wallA+wallBad = 4, L2 has wallB = 1
    l1 = _row(ws, DATA_START_ROW)
    l2 = _row(ws, DATA_START_ROW + 1)
    assert l1[1] == "L1" and l1[2] == 4
    assert l2[1] == "L2" and l2[2] == 1


def test_bad_group_by_falls_back_to_element_type() -> None:
    wb = _load(_elements(), group_by="nonsense")
    assert _row(wb["BOQ"], HEADER_ROW)[1] == "Element Type"


def test_detail_sheet_lists_every_element() -> None:
    wb = _load(_elements())
    ws = wb["Elements"]
    assert _row(ws, 1) == DETAIL_COLUMNS
    # 5 elements + 1 header
    assert ws.max_row == 6
    # rows sorted by (type, storey, name, stable_id); first is IfcBeam/L1
    first = _row(ws, 2)
    assert first[2] == "IfcBeam" and first[7] == 8.0  # element_type, length


def test_empty_elements_yields_zero_total() -> None:
    wb = _load([])
    ws = wb["BOQ"]
    # no group rows -> TOTAL lands on the first data row
    total = _row(ws, DATA_START_ROW)
    assert total[0] == "TOTAL"
    assert total[2] == 0


def test_unclassified_and_unassigned_fallbacks() -> None:
    els = [
        SimpleNamespace(
            id=uuid.uuid4(),
            stable_id="x",
            element_type=None,
            name="X",
            storey=None,
            discipline=None,
            quantities={},
        )
    ]
    wb = _load(els, group_by="element_type_storey")
    row = _row(wb["BOQ"], DATA_START_ROW)
    assert row[1] == "Unclassified"
    assert row[2] == "Unassigned"


def test_missing_quantities_blob_is_safe() -> None:
    els = [
        SimpleNamespace(
            id=uuid.uuid4(),
            stable_id="x",
            element_type="IfcWall",
            name="X",
            storey="L1",
            discipline="A",
            quantities=None,
        )
    ]
    wb = _load(els)
    row = _row(wb["BOQ"], DATA_START_ROW)
    assert row[2] == 1 and row[3] == 0.0  # count 1, area 0
