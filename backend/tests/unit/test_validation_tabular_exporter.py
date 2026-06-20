"""Unit tests for the validation CSV / XLSX exporters.

Pure tests - no DB import. Cover:
    1. row mapping: a report's results project to the documented flat shape.
    2. CSV formula-injection neutralisation: dangerous-leading cells are
       prefixed with an apostrophe in the emitted CSV (and XLSX).
    3. status / passed mapping and unicode survival.
"""

from __future__ import annotations

import csv
import io

from app.core.validation.engine import (
    RuleCategory,
    RuleResult,
    Severity,
    ValidationReport,
)
from app.modules.validation.tabular_exporter import (
    ROW_HEADERS,
    report_to_csv,
    report_to_rows,
    report_to_xlsx,
)

# ── Helpers ────────────────────────────────────────────────────────────────


def _r(
    rule_id: str,
    severity: Severity,
    *,
    passed: bool,
    message: str = "msg",
    element_ref: str | None = None,
    suggestion: str | None = None,
    category: RuleCategory = RuleCategory.COMPLIANCE,
) -> RuleResult:
    return RuleResult(
        rule_id=rule_id,
        rule_name=rule_id.replace(".", " ").title(),
        severity=severity,
        category=category,
        passed=passed,
        message=message,
        element_ref=element_ref,
        suggestion=suggestion,
    )


def _make_report(*results: RuleResult) -> ValidationReport:
    return ValidationReport(
        target_type="boq",
        target_id="boq-123",
        rule_sets_applied=["din276", "boq_quality"],
        results=list(results),
        duration_ms=4.2,
    )


def _parse_csv(blob: bytes) -> list[list[str]]:
    text = blob.decode("utf-8-sig")
    return list(csv.reader(io.StringIO(text)))


# ── 1. Row mapping ─────────────────────────────────────────────────────────


def test_report_to_rows_shape_and_order() -> None:
    """Headers match ROW_HEADERS and every result maps to one ordered row."""
    report = _make_report(
        _r("din276.kg_required", Severity.ERROR, passed=False, element_ref="pos-1", suggestion="add KG"),
        _r("boq_quality.zero_rate", Severity.WARNING, passed=False, message="rate is 0"),
        _r("boq_quality.no_dup", Severity.INFO, passed=True),
    )
    headers, rows = report_to_rows(report)

    assert headers == list(ROW_HEADERS)
    assert len(rows) == 3
    # Each row has exactly one cell per header.
    assert all(len(row) == len(ROW_HEADERS) for row in rows)

    col = {name: i for i, name in enumerate(headers)}
    first = rows[0]
    assert first[col["rule_id"]] == "din276.kg_required"
    # rule_name comes from the engine result verbatim; _r() title-cases the
    # rule id, and str.title() only breaks on '.'/space (not '_').
    assert first[col["rule_name"]] == "Din276 Kg_Required"
    assert first[col["severity"]] == "error"
    assert first[col["status"]] == "error"  # failing -> severity
    assert first[col["category"]] == "compliance"
    assert first[col["element_ref"]] == "pos-1"
    assert first[col["suggestion"]] == "add KG"


def test_status_passed_vs_failing() -> None:
    """A passing rule reads 'passed'; a failing one reads its severity."""
    report = _make_report(
        _r("a.passed", Severity.ERROR, passed=True),
        _r("a.warn", Severity.WARNING, passed=False),
        _r("a.info", Severity.INFO, passed=False),
    )
    headers, rows = report_to_rows(report)
    status_idx = headers.index("status")
    assert rows[0][status_idx] == "passed"
    assert rows[1][status_idx] == "warning"
    assert rows[2][status_idx] == "info"


def test_none_element_ref_and_suggestion_become_empty() -> None:
    """Missing element_ref / suggestion render as empty cells, not 'None'."""
    report = _make_report(_r("a.b", Severity.WARNING, passed=False))
    headers, rows = report_to_rows(report)
    col = {name: i for i, name in enumerate(headers)}
    assert rows[0][col["element_ref"]] == ""
    assert rows[0][col["suggestion"]] == ""


def test_empty_report_yields_no_rows() -> None:
    headers, rows = report_to_rows(_make_report())
    assert headers == list(ROW_HEADERS)
    assert rows == []


# ── 2. CSV formula-injection neutralisation ────────────────────────────────


def test_csv_neutralises_formula_in_message() -> None:
    """A message starting with '=' is written as literal text (apostrophe)."""
    payload = "=cmd|'/c calc'!A0"
    report = _make_report(_r("evil.rule", Severity.ERROR, passed=False, message=payload))
    rows = _parse_csv(report_to_csv(report))

    # Find the data row carrying the rule id.
    data_rows = [r for r in rows if r and r[0] == "evil.rule"]
    assert len(data_rows) == 1
    msg_idx = list(ROW_HEADERS).index("message")
    # csv module strips the outer quoting; the apostrophe guard must remain.
    assert data_rows[0][msg_idx] == "'" + payload
    assert data_rows[0][msg_idx].startswith("'=")


def test_csv_neutralises_all_dangerous_leads() -> None:
    """Every OWASP-listed dangerous lead char is neutralised in the cell."""
    for lead in ("=", "+", "-", "@"):
        payload = f"{lead}DANGER(1)"
        report = _make_report(
            _r("r.id", Severity.WARNING, passed=False, message=payload, suggestion=payload),
        )
        rows = _parse_csv(report_to_csv(report))
        data = [r for r in rows if r and r[0] == "r.id"][0]
        msg_idx = list(ROW_HEADERS).index("message")
        sug_idx = list(ROW_HEADERS).index("suggestion")
        assert data[msg_idx] == "'" + payload
        assert data[sug_idx] == "'" + payload


def test_csv_does_not_alter_safe_values() -> None:
    """A benign description is emitted unchanged (no spurious apostrophe)."""
    safe = "Concrete C30/37 missing rate"
    report = _make_report(_r("r.id", Severity.WARNING, passed=False, message=safe))
    rows = _parse_csv(report_to_csv(report))
    data = [r for r in rows if r and r[0] == "r.id"][0]
    msg_idx = list(ROW_HEADERS).index("message")
    assert data[msg_idx] == safe


def test_csv_header_present_and_unicode_survives() -> None:
    """The findings header row exists and non-ASCII messages round-trip."""
    report = _make_report(_r("r.id", Severity.ERROR, passed=False, message="Größe fehlt - Wänd€"))
    blob = report_to_csv(report)
    rows = _parse_csv(blob)
    assert list(ROW_HEADERS) in rows
    data = [r for r in rows if r and r[0] == "r.id"][0]
    assert "Größe fehlt - Wänd€" in data[list(ROW_HEADERS).index("message")]
    # UTF-8 BOM so Excel detects encoding.
    assert blob.startswith(b"\xef\xbb\xbf")


# ── 3. XLSX smoke + neutralisation ─────────────────────────────────────────


def test_xlsx_is_a_valid_workbook_with_rows() -> None:
    """The XLSX export opens with openpyxl and contains the findings."""
    from openpyxl import load_workbook

    report = _make_report(
        _r("din276.kg_required", Severity.ERROR, passed=False, element_ref="pos-1"),
        _r("boq_quality.zero_rate", Severity.WARNING, passed=False),
    )
    wb = load_workbook(io.BytesIO(report_to_xlsx(report)))
    ws = wb.active
    flat = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    assert "din276.kg_required" in flat
    assert "boq_quality.zero_rate" in flat
    # Header cells present.
    assert "rule_id" in flat
    assert "message" in flat


def test_xlsx_neutralises_formula() -> None:
    """A dangerous-leading message lands as literal text in the .xlsx cell."""
    from openpyxl import load_workbook

    payload = "=HYPERLINK(1)"
    report = _make_report(_r("evil", Severity.ERROR, passed=False, message=payload))
    wb = load_workbook(io.BytesIO(report_to_xlsx(report)))
    ws = wb.active
    values = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    assert ("'" + payload) in values
    # The raw, un-prefixed formula must NOT appear anywhere.
    assert payload not in values
